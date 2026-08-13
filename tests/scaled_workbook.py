"""A SYNTHETIC workbook shaped like the owner's shop, in the Test4 layout.

``tests/sample_workbook.py`` is 3 SO lines over 2 items on 6 machines — it plans
to **5 schedule entries**, which is enough to prove the wiring holds and far too
little for a violation check to have anything to see. This module builds the
bigger book: 60 SO lines over 24 items, 14 machines (5 CNC + 2 VMC on two
shifts, 7 single-shift benches), 19 operators across both shifts, an 8-step
routing family and a Thursday weekly off.

**IT IS NOT THE OWNER'S BOOK.** ``Test5.xlsx`` and friends are gitignored, so
nothing in this repo can be. Every number measured on it is a comparison of two
engines on the SAME book — which is what makes the comparison fair — and none of
them is a forecast for Anvitech.

Shared deliberately between ``tests/test_roster_end_to_end.py`` and
``scripts/roster_vs_new.py`` so the table posted to the owner and the test that
pins it can never drift onto two different books.

``seed`` varies the cycle times, the quantities and the delivery dates; the shop
and the routing family are fixed. Two seeds are two order books for one factory.
"""
import datetime
import io
import random

import openpyxl

PLAN_START = datetime.date(2026, 8, 12)

MACHINES = [
    ("CNC lathe", "CNC 1", 250, 19.5), ("CNC lathe", "CNC 2", 250, 19.5),
    ("CNC lathe", "CNC 3", 250, 19.5), ("CNC lathe", "CNC 4", 250, 19.5),
    ("CNC lathe", "CNC 5", 250, 19.5),
    ("Vertical Machining center", "VMC 1", 500, 19.5),
    ("Vertical Machining center", "VMC 2", 500, 19.5),
    ("Band saw cutting machine", "BS1", 100, 9.5),
    ("Manual Deburring", "MD1", 80, 9.5), ("Manual Deburring", "MD2", 80, 9.5),
    ("Manual Inspection", "MI1", 150, 9.5),
    ("Manual Inspection", "MI2", 150, 9.5),
    ("Manual Washing", "MW1", 80, 9.5), ("Manual Packing", "MPK1", 60, 9.5),
]

_FIRST, _SECOND = "First shift", "Second shift"
OPERATORS = [
    # Two-shift machining crew: every CNC/VMC is coverable on both shifts, and
    # the qualification sets OVERLAP — which is what gives an engine the chance
    # to hop somebody between machines, and the split check something to see.
    ("Anturam", "CNC1/CNC2", _FIRST), ("Sanjay", "CNC2/CNC3", _FIRST),
    ("Narayan", "CNC3/CNC4", _FIRST), ("Sidhu", "CNC4/CNC5", _FIRST),
    ("Ramesh", "CNC5/CNC1", _FIRST), ("Vikas", "VMC1/VMC2", _FIRST),
    ("Prakash", "VMC2/VMC1", _FIRST),
    ("Gopal", "CNC1/CNC2/CNC3", _SECOND), ("Mahesh", "CNC3/CNC4", _SECOND),
    ("Dinesh", "CNC4/CNC5", _SECOND), ("Suresh", "CNC5/CNC1", _SECOND),
    ("Kiran", "CNC2/CNC3", _SECOND), ("Ravi", "VMC1/VMC2", _SECOND),
    ("Ajay", "VMC2/VMC1", _SECOND),
    # Bench crew. Kept bench-only on purpose: a person qualified on both a CNC
    # and a bench is absorbed onto the CNC by the roster's matching, and the
    # benches would be left a thinner pool (the 2026-08-12 generalist finding —
    # no longer a deadlock after Task 10c, but still a shape worth not mixing
    # into a fixture whose job is to compare two engines).
    ("Bhaskar", "BS1/MD1", _FIRST), ("Chetan", "MD1/MD2/MW1", _FIRST),
    ("Deepak", "MI1/MI2", _FIRST), ("Eknath", "MW1/MPK1/MI2", _FIRST),
    ("Farid", "BS1/MD2/MPK1", _FIRST),
]

STEPS = [
    ("BANDSAW", 1.5, "BS1"),
    ("CNC FIRST SIDE", 4.0, "CNC1/CNC2/CNC3"),
    ("CNC SECOND SIDE", 3.5, "CNC3/CNC4/CNC5"),
    ("VMC FIRST SIDE", 5.0, "VMC1/VMC2"),
    ("DEBURING", 1.0, "MD1/MD2"),
    ("WASHING", 0.5, "MW1"),
    ("INSP", 0.8, "MI1/MI2"),
    ("PACKING", 0.4, "MPK1"),
]


def _set(ws, row, mapping):
    for ci, val in mapping.items():
        ws.cell(row=row, column=ci + 1, value=val)


def build_scaled_bytes(n_items=24, n_orders=60, start=PLAN_START, seed=7) -> bytes:
    """The workbook bytes. Same layout the loader reads for the real thing."""
    rng = random.Random(seed)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Machine master")
    ws.append(["Machine Type", "Machine No", "Hr Rate (Rs)", "Available Hrs/Day"])
    for row in MACHINES:
        ws.append(list(row))

    ws = wb.create_sheet("Operator & shift Master")
    ws.append(["Operator Name", "Preferred Machines", "Shift", None,
               "Shift", "Start", "End", "Notes"])
    shifts = [("First shift", "08:00", "19:00", ""),
              ("Second shift", "19:00", "05:00", "next day")]
    for i in range(max(len(OPERATORS), len(shifts))):
        row = [None] * 8
        if i < len(OPERATORS):
            row[0], row[1], row[2] = OPERATORS[i]
        if i < len(shifts):
            row[4], row[5], row[6], row[7] = shifts[i]
        ws.append(row)

    ws = wb.create_sheet("Weekly off & holiday master")
    ws.append(["Category", "Name", "Day / Date"])
    ws.append(["Weekly Off", "", "Every Thursday"])

    ws = wb.create_sheet("Item's process Master")
    _set(ws, 1, {5: "Order Type", 6: "Raw material data", 11: "Process sequence"})
    header = {0: "Master Sr NO", 1: "Customer", 2: "Item Description",
              3: "Item code", 6: "RM type", 10: "MOQ/Batch qty"}
    for p in range(len(STEPS)):
        base = 12 + p * 5
        header.update({base: f"Process {p+1}",
                       base + 1: f"Process {p+1} cycle time",
                       base + 2: f"Process {p+1} Total time",
                       base + 3: f"Process {p+1} Suggested M/c",
                       base + 4: f"Process {p+1} Allotted M/c"})
    _set(ws, 2, header)

    items = []
    for i in range(n_items):
        code = f"ITEM-{i:03d}"
        items.append(code)
        n_steps = 4 + (i % (len(STEPS) - 3))
        row = {0: i + 1, 1: "ALFA", 2: f"PART {i}", 3: code,
               6: "Dia 20 SS", 10: 50}
        for p in range(n_steps):
            name, cycle, machines = STEPS[p]
            base = 12 + p * 5
            cycle = round(cycle * (0.6 + 0.8 * rng.random()), 2)
            row.update({base: name, base + 1: cycle, base + 2: cycle,
                        base + 3: machines})
        _set(ws, 3 + i, row)

    ws = wb.create_sheet("Sales Order (SO) list")
    _set(ws, 1, {5: "SONo", 8: "Customer Name", 19: "Sales Item Code",
                 20: "Sales Item Name", 21: "SO Qty", 23: "SO Delivery Date",
                 24: "Remarks", 27: "Pend SO Qty"})
    for j in range(n_orders):
        code = items[j % n_items]
        qty = rng.choice([20, 40, 60, 100, 150])
        due = start + datetime.timedelta(days=3 + rng.randrange(16))
        _set(ws, 2 + j, {5: f"26-27SO{j:03d}", 8: "ALFA LAVAL", 19: code,
                         20: f"PART {j % n_items}", 21: qty, 23: due,
                         24: "", 27: qty})
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
