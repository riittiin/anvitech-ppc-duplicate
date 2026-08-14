"""One plan, through the REAL seam, audited by the shop's OWN rule checks.

Every other CP test drives one module. This one drives the whole chain the app
drives — ``load_all`` → the order book's SO lines → ``pipeline.run_forward``
with ``scheduler="cp"`` → ``engine/cp_adapter`` → ``cp_engine.decode`` → the
``ScheduleEntry`` list every surface publishes — and then hands the result to
``roster_engine.report``, an INDEPENDENT implementation of the same four shop
rules written for a different engine. That is what makes a green result mean
something: a model bug that the model's own checks share cannot hide here.

TWO SHOP SHAPES, AND THE DIFFERENCE BETWEEN THEM IS THE FINDING
---------------------------------------------------------------
Completion-DATE drift — the replayed plan finishing on a different day from the
solve that chose it — is **0 on a contended book and NOT an invariant**. Both
are measured here, deliberately, because a test suite that only ever showed the
zero would be publishing a guarantee this engine does not have:

  ``_contended_bytes()``   two CNCs and three benches, all two-shift, fought
                           over by five people, delivery dates tight enough that
                           the book is genuinely tardy. **Drift == 0, exactly,
                           no epsilon.** That is the assertion the plan exists
                           for, and it is asserted at zero.

  ``scaled_workbook``      the repo's own generated book, and the OWNER'S OWN
                           SHAPE: benches that run the day shift only (MD/MW/
                           MPK/MI/BS at 9.5 h/day) fed by CNC batches long
                           enough to run past 19:00. **It drifts, always LATE,
                           and it is RECORDED here rather than asserted away.**

Two distinct amplifiers were measured on that second shape, both rooted in the
one documented limitation — ``decode._JobState`` tracks one op at a time, so a
successor cannot be released while its predecessor is still in the chuck, while
the model's release is a linear bound on start variables that fires
mid-operation:

  * **the calendar amplifier, +1 day per order.** The bench step is released at
    19:30, the single-shift bench has already closed, and the next window is a
    day (or two, over the weekly off) away. This is Task 10's hand-built finding,
    reproduced here on the repo's own generated book rather than on a fixture
    built to show it.
  * **the ROSTER-GAP amplifier, +17 and +20 days**, measured 2026-08-14 on the
    loose book at a 30 s budget and new to this task. An op that slips past the
    shift the solve rostered its machine for lands in a shift ``cp_roster``
    deliberately left dark (``decode._Crew.staff``'s fourth case), so it waits
    for that machine's NEXT rostered shift — which on a book with slack can be
    weeks out. Read off the genome directly rather than inferred: VMC1's
    rostered shifts ran ``[… 12, 14, 15, 50, 96 …]``, the replay's VMC operation
    overran shift 15 at 05:00 on 22-08, and it resumed 11-09 — twenty days.
    ``decode._assign`` already documents this for a NEW order arriving after the
    solve; that it also catches a SOLVED order whose replay slipped is the part
    that was not known. **The magnitude is budget-dependent** — the same book at
    a 12 s budget returns a different genome and drifts +1/+1/+2 — so the two
    cases below are parametrized by the BOOK's tightness, never by a mechanism a
    given run has not been checked for.

**NEVER widen the tolerance to make the second shape pass.** Spec §5.3: tighten
the model, never loosen the decoder. The remedy is mid-shift re-staffing in the
decoder, named in ``decode.py``'s docstring and deliberately out of this plan's
scope.

THE SOLVE'S TIME LIMIT IS WALL CLOCK, so two runs of an identical book can
return different genomes: measured 2026-08-14, the scaled book at a 15 s budget
returned 49, then 67, then 58 late-days on three consecutive identical runs
(the same book at 20 s returned 43 five times out of five — converged, not
deterministic). Nothing below asserts a solved OBJECTIVE for that reason; what
is asserted is a property of the REPLAY, which holds for whichever plan the
search happens to return. ``ortools``' ``max_deterministic_time`` is the lever
if a genome ever has to reproduce across runs — see the Task 13 report.
"""
from __future__ import annotations

import datetime
import io
from dataclasses import replace

import pytest

pytest.importorskip("pyjobshop")

import openpyxl                                              # noqa: E402

from cp_engine import report as cp_report                    # noqa: E402
from engine import cp_adapter, pipeline                      # noqa: E402
from engine.config import Config                             # noqa: E402
from engine.loaders import load_all                          # noqa: E402
from engine.models import PlanRun                            # noqa: E402
import roster_engine.report as rr                            # noqa: E402
from tests.scaled_workbook import PLAN_START, build_scaled_bytes  # noqa: E402

# Seconds per solve. Sized from measurement, not guessed: at 8 s the scaled book
# comes back with no solution at all (``status_ok`` False, empty genome) and
# every drift assertion would pass vacuously — which is why ``_solved`` asserts
# the genome is non-empty before anything else looks at it. Every solve here is
# shared by a module-scoped fixture: three solves, not six, because this file
# adds real wall clock to a suite that runs in 87 s.
TIME_LIMIT = 12
BENCH_TIME_LIMIT = {0: 12, 10: 18}       # due_shift -> seconds; see the fixture


# --------------------------------------------------------------------------- #
# A CONTENDED book, built the way the app receives one: as a workbook
# --------------------------------------------------------------------------- #

_F, _S = "First shift", "Second shift"

# Benches are TWO-SHIFT here, and that is the one deliberate difference from
# ``tests/scaled_workbook``. It is not a convenience: it removes the calendar
# amplifier documented above, so this fixture measures replay fidelity and
# nothing else. The owner's single-shift-bench shape is measured too — on the
# scaled book, below — rather than being engineered out of existence.
_MACHINES = [
    ("CNC lathe", "CNC 1", 250, 19.5),
    ("CNC lathe", "CNC 2", 250, 19.5),
    ("Manual Deburring", "MD1", 80, 19.5),
    ("Manual Deburring", "MD2", 80, 19.5),
    ("Manual Inspection", "MI1", 150, 19.5),
]
# Five people over five machines with OVERLAPPING qualifications, one of them
# alone on the second-shift benches: that is what makes the per-shift roster a
# real decision rather than a formality.
_OPERATORS = [
    ("Anturam", "CNC1/CNC2", _F), ("Gopal", "CNC1/CNC2", _S),
    ("Chetan", "MD1/MD2/MI1", _F), ("Deepak", "MD1/MD2/MI1", _F),
    ("Farid", "MD1/MD2/MI1", _S),
]
_STEPS = [("CNC FIRST SIDE", 2.0, "CNC1/CNC2"), ("DEBURING", 0.5, "MD1/MD2"),
          ("INSP", 0.3, "MI1")]


def _set(ws, row, mapping):
    for ci, val in mapping.items():
        ws.cell(row=row, column=ci + 1, value=val)


def _contended_bytes(n_orders=14, start=PLAN_START, seed=7) -> bytes:
    """A tardy book on a contended shop, in the Test4 layout the loader reads.

    Written as a WORKBOOK rather than as hand-built ``Batch`` objects on
    purpose: every other CP test starts downstream of the loader, and the seam
    this file exists to prove includes it.
    """
    import random
    rng = random.Random(seed)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Machine master")
    ws.append(["Machine Type", "Machine No", "Hr Rate (Rs)", "Available Hrs/Day"])
    for row in _MACHINES:
        ws.append(list(row))

    ws = wb.create_sheet("Operator & shift Master")
    ws.append(["Operator Name", "Preferred Machines", "Shift", None,
               "Shift", "Start", "End", "Notes"])
    shifts = [("First shift", "08:00", "19:00", ""),
              ("Second shift", "19:00", "05:00", "next day")]
    for i in range(max(len(_OPERATORS), len(shifts))):
        row = [None] * 8
        if i < len(_OPERATORS):
            row[0], row[1], row[2] = _OPERATORS[i]
        if i < len(shifts):
            row[4], row[5], row[6], row[7] = shifts[i]
        ws.append(row)

    ws = wb.create_sheet("Weekly off & holiday master")
    ws.append(["Category", "Name", "Day / Date"])
    ws.append(["Weekly Off", "", "Every Thursday"])

    ws = wb.create_sheet("Item's process Master")
    _set(ws, 1, {5: "Order Type", 6: "Raw material data", 11: "Process sequence"})
    header = {0: "Master Sr NO", 1: "Customer", 2: "Item Description",
              3: "Item code", 6: "RM type", 10: "MOQ/Batch qty"}
    for p in range(len(_STEPS)):
        base = 12 + p * 5
        header.update({base: f"Process {p+1}",
                       base + 1: f"Process {p+1} cycle time",
                       base + 2: f"Process {p+1} Total time",
                       base + 3: f"Process {p+1} Suggested M/c",
                       base + 4: f"Process {p+1} Allotted M/c"})
    _set(ws, 2, header)

    items = [f"IT{i:02d}" for i in range(n_orders)]
    for i, code in enumerate(items):
        row = {0: i + 1, 1: "ALFA", 2: f"PART {i}", 3: code, 6: "Dia 20 SS",
               10: 50}
        for p, (name, cycle, machines) in enumerate(_STEPS):
            base = 12 + p * 5
            row.update({base: name, base + 1: cycle, base + 2: cycle,
                        base + 3: machines})
        _set(ws, 3 + i, row)

    ws = wb.create_sheet("Sales Order (SO) list")
    _set(ws, 1, {5: "SONo", 8: "Customer Name", 19: "Sales Item Code",
                 20: "Sales Item Name", 21: "SO Qty", 23: "SO Delivery Date",
                 24: "Remarks", 27: "Pend SO Qty"})
    for j in range(n_orders):
        qty = rng.choice([80, 120, 160, 200])
        due = start + datetime.timedelta(days=1 + rng.randrange(3))
        _set(ws, 2 + j, {5: f"26-27SO{j:03d}", 8: "ALFA LAVAL", 19: items[j],
                         20: f"PART {j}", 21: qty, 23: due, 24: "", 27: qty})
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load(raw, *, due_shift=0):
    """``(so_lines, masters)`` from workbook bytes.

    ``due_shift`` pulls every delivery date that many days EARLIER, in memory
    only. It is how the scaled book is made genuinely tardy: measured
    2026-08-14, that book as generated finishes entirely on time, and a solve on
    a zero-tardiness book stops at the first feasible schedule it finds — it
    measures FEASIBILITY and reports it as optimisation (findings §5). Nothing
    is written anywhere; ``load_all`` reads the bytes read-only.
    """
    so_lines, masters = load_all(io.BytesIO(raw))
    if due_shift:
        for line in so_lines:
            if line.delivery_date:
                line.delivery_date -= datetime.timedelta(days=due_shift)
    return so_lines, masters


def _config(**kw):
    base = dict(plan_start_date=PLAN_START, scheduler="cp", setup_time_min=90.0,
                apply_operator_logic=True, cp_time_limit_sec=TIME_LIMIT)
    base.update(kw)
    return Config(**base)


def _replay(so_lines, masters, config, genome=None):
    """The app's own plan path: Rules 1→2→3 then the cp seam, one call."""
    run = PlanRun(so_lines=list(so_lines))
    cfg = replace(config, cp_genome=genome) if genome is not None else config
    trace = pipeline.run_forward(run, cfg, masters)
    assert trace["rule6"]["error"] is None, trace["rule6"]["error"]
    return run, trace


def _solved(so_lines, masters, config):
    """Solve through the seam and return the result, asserting it is REAL.

    A solve that times out before finding anything returns an ``OptimizeResult``
    with no genome, and ``completion_drift`` on an empty genome returns ``[]``
    by design (it has nothing to compare). Every drift assertion below would
    then pass while measuring nothing at all, so the genome is checked here,
    once, for all of them.
    """
    result = cp_adapter.solve(list(so_lines), config, masters, budget_evals=0,
                              seed=42)
    assert result.genome, "the solve returned no plan — raise the time limit"
    assert result.genome.get("cp_completion"), "genome carries no completions"
    return result


class _Run:
    """One solve and its replay, shared by every test that needs them."""

    def __init__(self, raw, config, due_shift=0):
        self.so_lines, self.masters = _load(raw, due_shift=due_shift)
        self.config = config
        self.due_shift = due_shift
        self.result = _solved(self.so_lines, self.masters, config)
        self.run, _trace = _replay(self.so_lines, self.masters, config,
                                   genome=self.result.genome)
        self.entries = self.run.schedule
        self.batches = self.run.batches_prioritized
        self.genome = self.result.genome


@pytest.fixture(scope="module")
def contended():
    """The contended book, solved and replayed once for the whole module."""
    return _Run(_contended_bytes(), _config())


@pytest.fixture(scope="module")
def bench_shape(request):
    """The owner's single-shift-bench shape, at one due-date tightness."""
    due_shift = request.param
    return _Run(build_scaled_bytes(n_items=12, n_orders=20, seed=7),
                _config(cp_time_limit_sec=BENCH_TIME_LIMIT[due_shift]),
                due_shift=due_shift)


# --------------------------------------------------------------------------- #
# 1. A plan with NO genome at all still breaks none of the four rules
# --------------------------------------------------------------------------- #

def test_a_genome_less_cp_plan_breaks_none_of_the_four_rules():
    """The honest worst case: the decoder falls back for EVERY operation.

    This is the state of the world after an upload and before the first search,
    and it is the state Render is in on a cold store — so the rules have to hold
    without a genome, not only with one.
    """
    so_lines, masters = _load(build_scaled_bytes(n_items=12, n_orders=20, seed=7))
    config = _config()
    run, _trace = _replay(so_lines, masters, config)
    assert run.schedule

    counts: dict = {}
    for row in rr.all_violations(run.schedule, masters, config,
                                 batches=run.batches_prioritized):
        counts[row["kind"]] = counts.get(row["kind"], 0) + 1
    for kind in cp_report.RULE_KINDS:
        assert counts.get(kind, 0) == 0, (kind, counts)
    # Nothing may be silently missing from the plan either: an order in no plan
    # at all is a different failure from an order planned badly, and it has its
    # own check. It is asserted here because a genome-less replay is exactly
    # where a fallback could quietly drop a job.
    assert counts.get("UNPLANNED_ORDER", 0) == 0, counts


# --------------------------------------------------------------------------- #
# 2. THE ASSERTION THIS TASK EXISTS FOR — zero drift on a contended book
# --------------------------------------------------------------------------- #

def test_a_solved_contended_book_replays_with_ZERO_date_drift(contended, capsys):
    """Spec §8: on the book that was solved, the published completion DATE must
    be the solved one. Exactly — ``== 0``, no epsilon, no tolerance.

    A non-zero here means the plan on screen is not the plan that was solved, so
    the late-days the search minimised are not the late-days the shop realises.
    """
    assert cp_report.completion_drift(contended.entries, contended.genome) == []

    # NON-VACUITY, three ways: the book really was tardy (a zero-tardiness book
    # is the easy case and proves nothing), every batch really was compared, and
    # the genome really is the one that produced these entries.
    best = contended.result.best or {}
    assert best.get("cp_total_late_days", 0) > 0, best
    solved_keys = set(contended.genome["cp_completion"])
    assert solved_keys == {e.batch_id for e in contended.entries}
    assert len(solved_keys) == len(contended.batches)
    assert cp_report.genome_stale(contended.batches, contended.masters,
                                  contended.genome) == []
    with capsys.disabled():
        print(f"\n  contended book: {len(solved_keys)} batches, "
              f"{best.get('cp_total_late_days')} late-days "
              f"({best.get('cp_status')}), date drift 0")


def test_the_four_rule_checks_are_clean_on_the_SOLVED_replay(contended, capsys):
    """The same plan, audited by the four checks written for the other engine.

    Partitioned on the row's own ``breach`` key rather than on an out-of-band
    list of kinds: ``IDLE_CAPACITY`` is a capacity MEASUREMENT and is
    legitimately non-zero under this engine (E1 forbids an operation spanning an
    unmanned shift, and the decoder defers a pool-staffed machine's fallback
    work by one shift on purpose), so asserting the undifferentiated list empty
    would be asserting something false.
    """
    breaches, measured = cp_adapter.plan_violations(
        contended.entries, contended.masters, contended.config,
        batches=contended.batches, genome=contended.genome)
    assert breaches == [], [r["message"] for r in breaches]
    assert all(r["kind"] == "IDLE_CAPACITY" for r in measured), measured
    with capsys.disabled():
        print(f"  four rule checks on the solved replay: 0 breaches, "
              f"{len(measured)} capacity measurement(s)")


def test_the_rule_checks_would_CATCH_a_breach_in_this_very_plan(contended):
    """Non-vacuity for the two tests above, and it is not optional.

    "0 breaches" is exactly what a deleted check reports. So the real plan is
    corrupted here in the one way the shop cares about — the same machine
    running two operations at once — and the check must find it. Without this,
    ripping ``machine_conflict_violations`` out entirely would leave the suite
    green.
    """
    masters, config = contended.masters, contended.config
    on_machine = [e for e in contended.entries
                  if e.machine not in (cp_adapter.OS_LANE, cp_adapter.OFF_LANE)]
    assert len(on_machine) >= 2
    victim, clash = on_machine[0], on_machine[1]
    forged = replace(clash, machine=victim.machine, start=victim.start,
                     end=victim.end,
                     op_segments=[(victim.start, victim.end, clash.operator)])
    breaches, _measured = cp_adapter.plan_violations(
        list(contended.entries) + [forged], masters, config)
    kinds = {r["kind"] for r in breaches}
    assert "MACHINE_DOUBLE_BOOKED" in kinds, kinds


# --------------------------------------------------------------------------- #
# 3. THE SHAPE THAT DRIFTS — recorded, never asserted away
# --------------------------------------------------------------------------- #

# The two tightnesses are labelled by the BOOK, not by a mechanism: which
# amplifier dominates a given run depends on the genome the wall-clock-limited
# search happened to return, and naming a cause a particular run did not have
# checked is the thing this repo's reports are forbidden to do (2026-08-09).
# Both amplifiers are described, with their evidence, in the module docstring.
_TIGHTNESS = {
    10: "delivery dates pulled in 10 days (genuinely tardy)",
    0: "delivery dates as generated (loose)",
}


@pytest.mark.parametrize("bench_shape", [10, 0], indirect=True)
def test_the_owners_single_shift_bench_shape_DRIFTS_and_it_is_recorded(
        bench_shape, capsys):
    """Completion-DATE drift is 0 on a contended book and NOT an invariant.

    This is the owner's own shop shape — benches on the day shift only, fed by
    CNC batches that run past 19:00 — on the repo's own generated book, so it is
    not a fixture built to fail. What is asserted is what is TRUE and what
    matters: drift exists, and it is one-sided **LATE**. Late is conservative
    for the floor (work arrives earlier than the sheet says). EARLY would mean
    the decoder handed itself capacity the solver withheld, which is a defect,
    and that is the direction this test makes fail loudly.

    The magnitudes are PRINTED, never asserted: they are wall-clock dependent
    and nothing establishes a ceiling. Observed 2026-08-14 on this book: +1 day
    on 2-8 of 13 batches with the dates pulled in, and +17 / +20 days on two
    batches with them left loose.
    """
    rows = cp_report.completion_drift(bench_shape.entries, bench_shape.genome)
    assert rows, ("this book no longer drifts. That is GOOD NEWS, not a broken "
                  "test — read this file's module docstring, confirm the "
                  "decoder's concurrency was fixed, and delete this case rather "
                  "than weakening the zero-drift assertion above.")
    for row in rows:
        assert row["days"] > 0, row          # LATE, never EARLY
        assert row["solved"] in row["message"]
        assert row["replayed"] in row["message"]
    # The genome is not stale — the drift is the replay's, not a moved book.
    assert cp_report.genome_stale(bench_shape.batches, bench_shape.masters,
                                  bench_shape.genome) == []
    with capsys.disabled():
        print(f"\n  scaled book, {_TIGHTNESS[bench_shape.due_shift]}:\n    "
              + ", ".join(f"{r['batch_id']} {r['solved']}->{r['replayed']} "
                          f"(+{r['days']}d)" for r in rows[:8]))


# --------------------------------------------------------------------------- #
# 4. THE WORKER PATH, DRIVEN — not read
# --------------------------------------------------------------------------- #

def test_the_cloud_worker_round_trip_carries_a_REAL_genome_home(contended,
                                                                capsys):
    """The whole off-box hop, with an actual solve in the middle.

    ``tests/test_cp_wiring.py`` pins every joint of this path with the solve
    monkeypatched, which proves the plumbing but not that a real genome survives
    it. This drives it: ``build_payload`` (JSON) → ``parse_payload`` →
    ``run_candidate`` (a REAL CP solve, in the worker's own entry point) →
    ``merge_shard_rows`` (the reduce the sharded workflow performs) →
    ``genome_of_winner`` (how the app recovers decisions from the stripped table
    the non-sharded worker posts back) → the app's replay. The recovered genome
    must reproduce the recovered ranks' plan with ZERO drift, because that
    combination is exactly what Apply stores and ``/run`` replays.

    What this does NOT cover, and no test in this repo can: GitHub Actions
    dispatch, the Oracle claim window, and the HTTP hop itself. Those are
    verified by code path only — see the Task 13 report.
    """
    from engine import optimize_service
    from engine.models import Order

    orders = {}
    for line in contended.so_lines:
        o = Order(line.so_no, line.item_code, line.item_name, line.qty,
                  line.delivery_date)
        orders[o.key] = o
    cfg = contended.config
    payload = optimize_service.build_payload(
        orders, [], _contended_bytes(), cfg, seed=42,
        candidates=optimize_service.cloud_candidates(cfg),
        budget_per_candidate=optimize_service.cloud_budget(cfg))
    # JSON-safety is the point of a payload: the real hop is an HTTP body.
    import json
    payload = json.loads(json.dumps(payload, default=str))

    jobs = optimize_service.contest_jobs(payload)
    assert len(jobs) == 1, jobs        # a CP contest is ONE solve, not a sweep
    overlap, flexible, seed = jobs[0]
    row = optimize_service.run_candidate(payload, overlap, flexible, seed=seed)
    assert row["genome"], "the genome never left run_candidate"

    merged = optimize_service.merge_shard_rows(payload, [row], row["evals"],
                                               False)
    recovered = optimize_service.genome_of_winner(
        merged["rows"], merged["winner_overlap"], merged["winner_flexible"],
        merged["best"])
    assert recovered == row["genome"]

    run, _trace = _replay(contended.so_lines, contended.masters, cfg,
                          genome=recovered)
    assert cp_report.completion_drift(run.schedule, recovered) == []
    assert merged["ranks"], "no job order came home"
    with capsys.disabled():
        print(f"  cloud round trip: {len(recovered)} genome keys, "
              f"{len(merged['ranks'])} ranks, drift 0")


def test_a_deep_search_with_no_solver_fails_in_the_OWNERS_words():
    """cp has NO LOCAL SEARCH FALLBACK, and that is a documented limitation.

    It cannot be otherwise: the solver is deliberately absent from
    requirements.txt so it can never reach the app server (which only replays a
    finished plan), and a 15-minute CP solve on a 0.1-CPU free instance would be
    a fallback nobody wants even if it were installed — cores buy the bound, and
    that box has a fraction of one.

    What WAS a defect is how it read. The raw ``ImportError`` surfaces verbatim
    in ``_OPTIMIZE["error"]`` and in the Done button's durable note, so the floor
    would have been told **"No module named 'ortools'"** — visible but
    unactionable, which is the 2026-08-09 class exactly. The message must say
    which of two things happened (the search failed / the plan is unchanged) and
    what to do next.

    This test also pins WHERE the guard has to sit, which a first attempt got
    wrong: ``cp_engine.solve`` imports pyjobshop and ortools inside its own
    functions, so ``from cp_engine import solve`` SUCCEEDS on a box with neither
    and the error comes from deep inside the solve. A guard on the import line
    passed review by eye and never fired.
    """
    import sys

    class _Blocker:
        def find_spec(self, name, path=None, target=None):
            if name.split(".")[0] in ("pyjobshop", "ortools"):
                raise ImportError("blocked: " + name)
            return None

    saved = {k: v for k, v in sys.modules.items()
             if k.split(".")[0] in ("pyjobshop", "ortools")
             or k == "cp_engine.solve"}
    for key in saved:
        del sys.modules[key]
    sys.meta_path.insert(0, _Blocker())
    try:
        so_lines, masters = _load(_contended_bytes(n_orders=2))
        with pytest.raises(RuntimeError) as err:
            cp_adapter.solve(list(so_lines), _config(), masters, budget_evals=0)
    finally:
        sys.meta_path.pop(0)
        sys.modules.update(saved)
    text = str(err.value)
    assert "pyjobshop" not in text, text     # not the library's words
    assert "unchanged" in text, text         # the plan on screen is safe
    assert "worker" in text, text            # and what to do about it


def test_the_solve_is_given_the_configured_number_of_CORES():
    """**Cores buy the bound.** Measured (findings §1): 2 → 4 workers at the same
    30 minutes moved the proven floor 170 → 215 late-days and cut the gap from
    2.41× to 1.99×, while six times the wall clock at 2 workers moved it by two
    days. It was found ON THIS TASK that ``cp_adapter.solve`` read the count off
    a config field that did not exist, so ``or 1`` won and **every solve ran
    single-threaded** — the one measured lever, off, with no error and nothing on
    screen. The deployment note ("pin the worker at 4+ cores") is worthless if
    the solver is told to use one, so the field and the default are pinned here.
    """
    import types
    from cp_engine import solve as cp_solve

    seen = {}

    def _spy(*args, **kw):
        seen.update(kw)
        return types.SimpleNamespace(status_ok=False, genome={}, status="X",
                                     total_late_days=0, spread=0,
                                     lower_bound_days=0)

    so_lines, masters = _load(_contended_bytes(n_orders=2))
    original = cp_solve.solve_book
    cp_solve.solve_book = _spy
    try:
        cp_adapter.solve(list(so_lines), _config(cp_num_workers=6), masters,
                         budget_evals=0, seed=42)
    finally:
        cp_solve.solve_book = original
    assert seen.get("num_workers") == 6, seen
    assert Config().cp_num_workers >= 4, "the measured shipping default is 4+"


# --------------------------------------------------------------------------- #
# 5. The published plan is a plan, not just a rule-legal set of intervals
# --------------------------------------------------------------------------- #

def test_every_order_in_the_book_reaches_the_published_plan(contended):
    """A rule-legal plan that quietly drops orders passes every check above.

    This is the 2026-08-11 class stated as an invariant: the plan must cover the
    batches Rule 1 produced, and every real-machine entry must name an operator
    and carry a segment (``engine/freeze.py`` pins machine AND operator, so an
    empty name freezes a ghost).
    """
    planned = {e.batch_id for e in contended.entries}
    assert planned == {str(b.batch_id) for b in contended.batches}
    for entry in contended.entries:
        if entry.machine in (cp_adapter.OS_LANE, cp_adapter.OFF_LANE):
            continue
        assert entry.operator, entry
        assert entry.op_segments, entry
        assert entry.end >= entry.start
